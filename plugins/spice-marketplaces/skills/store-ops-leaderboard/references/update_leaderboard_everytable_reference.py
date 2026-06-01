"""
Everytable Store Leaderboard — Weekly Update Script
====================================================
Rotates current Data Input → Previous Week, loads new platform exports,
updates period label. Leaderboard formulas handle the rest.

USAGE:
  1. Update the CONFIG section below with this week's date range and file paths.
  2. Run: python update_leaderboard.py
  3. Open the workbook in Excel / Sheets to verify rankings.

REQUIRED EXPORTS:
  - UE: Order History CSV + Restaurant Ratings CSV (from UE Manager)
  - DD: Financial Transactions CSV + Cancelled Orders CSV + Missing/Incorrect CSV
  - DD (optional): Customer Reviews CSV (for "Loved" ratings)
  - GH: Operations Review CSV (from GH for Business)
"""
import csv
import sys
from datetime import datetime
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

# ╔══════════════════════════════════════════════╗
# ║  CONFIG — UPDATE THESE EACH WEEK            ║
# ╚══════════════════════════════════════════════╝

PERIOD_START = datetime(2026, 1, 1)
PERIOD_END = datetime(2026, 3, 31)
PERIOD_LABEL = "Q1 2026 (Jan-Mar)"

# File paths — update to match this week's uploaded exports
UE_ORDERS = ""       # UE Order History CSV
UE_RATINGS = ""      # UE Restaurant Ratings CSV
DD_FIN = ""          # DD Financial Transactions CSV
DD_CANCELS = ""      # DD Cancelled Orders CSV
DD_ERRORS = ""       # DD Missing/Incorrect Orders CSV
DD_REVIEWS = ""      # DD Customer Reviews CSV (optional, set "" to skip)
GH_OPS = ""          # GH Operations Review CSV

WORKBOOK = "Everytable_Store_Leaderboard.xlsx"

# Stores to exclude (confirmed closed or inactive)
EXCLUDE = {'CA062'}  # Anaheim
MIN_ORDERS = 5       # Stores with fewer orders get excluded

# ╔══════════════════════════════════════════════╗
# ║  STYLES                                      ║
# ╚══════════════════════════════════════════════╝

LIGHT_GRAY = "F5F5F5"
dat_font = Font(name="Arial", size=10)
thin_b = Border(
    left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))

# ╔══════════════════════════════════════════════╗
# ║  HELPERS                                     ║
# ╚══════════════════════════════════════════════╝

def parse_date(d):
    """Parse date from CSV. Handles 'YYYY-MM-DD' and 'YYYY-MM-DD HH:MM:SS' formats."""
    return datetime.strptime(d.strip().split(' ')[0], '%Y-%m-%d')

def in_period(d):
    return PERIOD_START <= d <= PERIOD_END

def dd_store_id(name):
    """Extract store ID (e.g., CA001) from DD store name like 'CA001 - Everytable - Location'."""
    return name.split(' - ')[0].strip() if ' - ' in name else name.split(' ')[0]

def validate_files():
    """Check that all required files are set and exist."""
    required = {
        'UE Orders': UE_ORDERS,
        'UE Ratings': UE_RATINGS,
        'DD Financial': DD_FIN,
        'DD Cancels': DD_CANCELS,
        'DD Errors': DD_ERRORS,
        'GH Operations': GH_OPS,
    }
    missing = [name for name, path in required.items() if not path]
    if missing:
        print(f"ERROR: Missing file paths for: {', '.join(missing)}")
        print("Update the CONFIG section at the top of this script.")
        sys.exit(1)

    import os
    not_found = [name for name, path in required.items() if path and not os.path.exists(path)]
    if not_found:
        print(f"ERROR: Files not found: {', '.join(not_found)}")
        for name, path in required.items():
            if name in not_found:
                print(f"  {name}: {path}")
        sys.exit(1)

    if DD_REVIEWS:
        if not os.path.exists(DD_REVIEWS):
            print(f"WARNING: DD Reviews file not found: {DD_REVIEWS}")
            print("  DD 'Loved' ratings will be 0. Continuing without it.")

# ╔══════════════════════════════════════════════╗
# ║  MAIN                                        ║
# ╚══════════════════════════════════════════════╝

def main():
    validate_files()

    # ── BUILD STORE ID → NAME MAP ──
    with open(UE_ORDERS, encoding='utf-8') as f:
        ue_rows = list(csv.DictReader(f))

    id_to_name = {}
    for r in ue_rows:
        sid = r['ID de tienda externa']
        name = r['Tienda'].replace("Everytable (", "").rstrip(")")
        id_to_name[sid] = name

    with open(GH_OPS, encoding='utf-8') as f:
        gh_rows = list(csv.DictReader(f))
    for r in gh_rows:
        sid = r['store_number']
        if sid not in id_to_name:
            id_to_name[sid] = r['city']

    print(f"Store map: {len(id_to_name)} stores")

    # ── 1. UBER EATS ──
    ue_data = defaultdict(lambda: {'orders': 0, 'cancels': 0, 'errors': 0, 'ratings': 0})
    for r in ue_rows:
        d = parse_date(r['Fecha del pedido'])
        if not in_period(d): continue
        sid = r['ID de tienda externa']
        if sid in EXCLUDE: continue
        ue_data[sid]['orders'] += 1
        status = r['Estado del pedido']
        if status == 'canceled': ue_data[sid]['cancels'] += 1
        elif status in ('failed', 'unfulfilled'): ue_data[sid]['errors'] += 1

    with open(UE_RATINGS, encoding='utf-8') as f:
        for r in csv.DictReader(f):
            d = parse_date(r['Fecha del pedido'])
            if not in_period(d): continue
            sid = r['ID de tienda externa']
            if sid in EXCLUDE: continue
            if r['Valor de la valoración'] == '5':
                ue_data[sid]['ratings'] += 1

    print(f"UE: {len(ue_data)} stores, {sum(d['orders'] for d in ue_data.values())} orders, "
          f"{sum(d['ratings'] for d in ue_data.values())} ratings")

    # ── 2. DOORDASH ──
    dd_data = defaultdict(lambda: {'orders': 0, 'cancels': 0, 'errors': 0, 'ratings': 0})

    with open(DD_FIN, encoding='utf-8') as f:
        for r in csv.DictReader(f):
            if r['Tipo de transacción'] != 'Order': continue
            d = parse_date(r['Hora local del sello de tiempo'])
            if not in_period(d): continue
            sid = dd_store_id(r['Nombre de la tienda'])
            if sid in EXCLUDE: continue
            dd_data[sid]['orders'] += 1

    with open(DD_CANCELS, encoding='utf-8') as f:
        for r in csv.DictReader(f):
            d = parse_date(r['Order Placed Date'])
            if not in_period(d): continue
            sid = dd_store_id(r['Store Name'])
            if sid in EXCLUDE: continue
            dd_data[sid]['cancels'] += 1

    seen_orders = defaultdict(set)
    with open(DD_ERRORS, encoding='utf-8') as f:
        for r in csv.DictReader(f):
            d = parse_date(r['Order Delivered Date'])
            if not in_period(d): continue
            sid = dd_store_id(r['Store Name'])
            if sid in EXCLUDE: continue
            oid = r['DD Order ID']
            if oid not in seen_orders[sid]:
                seen_orders[sid].add(oid)
                dd_data[sid]['errors'] += 1

    # DD Reviews (optional)
    if DD_REVIEWS:
        import os
        if os.path.exists(DD_REVIEWS):
            # TODO: Parse DD Customer Reviews CSV for "Loved" counts
            # Format TBD — need a sample export to confirm headers
            print("  DD Reviews file provided — parsing not yet implemented (need sample format)")
        else:
            print("  DD Reviews file not found, skipping")

    print(f"DD: {len(dd_data)} stores, {sum(d['orders'] for d in dd_data.values())} orders, "
          f"{sum(d['cancels'] for d in dd_data.values())} cancels, "
          f"{sum(d['errors'] for d in dd_data.values())} errors")
    if not DD_REVIEWS:
        print("  ⚠️  DD 'Loved' ratings = 0 (no Customer Reviews export)")

    # ── 3. GRUBHUB ──
    gh_data = {}
    for r in gh_rows:
        sid = r['store_number']
        if sid in EXCLUDE: continue
        gh_data[sid] = {
            'orders': int(r['total_orders'] or 0),
            'cancels': int(r['total_canceled_orders'] or 0),
            'errors': int(r['adjusted_orders'] or 0),
            'ratings': int(r['ratings_5_stars'] or 0),
        }
    print(f"GH: {len(gh_data)} stores, {sum(d['orders'] for d in gh_data.values())} orders, "
          f"{sum(d['ratings'] for d in gh_data.values())} ratings")

    # ── 4. CANONICAL STORE LIST ──
    all_sids = sorted(
        (set(ue_data.keys()) | set(dd_data.keys()) | set(gh_data.keys())) - EXCLUDE
    )
    # Filter low-activity stores
    all_sids = [s for s in all_sids if (
        ue_data.get(s, {}).get('orders', 0) +
        dd_data.get(s, {}).get('orders', 0) +
        gh_data.get(s, {}).get('orders', 0)
    ) >= MIN_ORDERS]

    store_names = [id_to_name.get(sid, sid) for sid in all_sids]
    print(f"\nCanonical stores: {len(store_names)}")

    # ── 5. LOAD WORKBOOK & ROTATE ──
    wb = load_workbook(WORKBOOK)

    # Rotate: Copy current Data Input → Previous Week
    ws_data = wb['Data Input']
    ws_prev = wb['Previous Week']

    # Clear Previous Week
    for r in range(2, 500):
        for col in range(1, 7):
            ws_prev.cell(row=r, column=col).value = None

    # Copy current Data Input to Previous Week
    copied = 0
    for r in range(2, 500):
        val = ws_data.cell(row=r, column=1).value
        if not val: break
        for col in range(1, 7):
            c_src = ws_data.cell(row=r, column=col)
            c_dst = ws_prev.cell(row=r, column=col, value=c_src.value)
            c_dst.font = dat_font
            c_dst.border = thin_b
            if r % 2 == 0: c_dst.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        copied += 1
    print(f"Rotated {copied} rows to Previous Week")

    # ── 6. UPDATE STORES TAB ──
    ws_stores = wb['Stores']
    for r in range(2, 60):
        ws_stores.cell(row=r, column=1).value = None
    for i, name in enumerate(store_names):
        c = ws_stores.cell(row=i+2, column=1, value=name)
        c.font = dat_font; c.border = thin_b

    # ── 7. CLEAR & LOAD DATA INPUT ──
    for r in range(2, 500):
        for col in range(1, 7):
            ws_data.cell(row=r, column=col).value = None

    row = 2
    for sid, name in zip(all_sids, store_names):
        # UE row
        ue = ue_data.get(sid, {'orders': 0, 'cancels': 0, 'errors': 0, 'ratings': 0})
        if ue['orders'] > 0:
            for col, val in enumerate([name, "Uber Eats", ue['orders'], ue['ratings'], ue['cancels'], ue['errors']], 1):
                c = ws_data.cell(row=row, column=col, value=val)
                c.font = dat_font; c.border = thin_b
                if row % 2 == 0: c.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
            row += 1

        # DD row
        dd = dd_data.get(sid, {'orders': 0, 'cancels': 0, 'errors': 0, 'ratings': 0})
        if dd['orders'] > 0:
            for col, val in enumerate([name, "DoorDash", dd['orders'], dd['ratings'], dd['cancels'], dd['errors']], 1):
                c = ws_data.cell(row=row, column=col, value=val)
                c.font = dat_font; c.border = thin_b
                if row % 2 == 0: c.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
            row += 1

        # GH row
        gh = gh_data.get(sid, {'orders': 0, 'cancels': 0, 'errors': 0, 'ratings': 0})
        if gh['orders'] > 0:
            for col, val in enumerate([name, "Grubhub", gh['orders'], gh['ratings'], gh['cancels'], gh['errors']], 1):
                c = ws_data.cell(row=row, column=col, value=val)
                c.font = dat_font; c.border = thin_b
                if row % 2 == 0: c.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
            row += 1

    data_rows = row - 2
    print(f"Data Input: {data_rows} rows loaded")

    # ── 8. UPDATE PERIOD LABEL ──
    ws_lb = wb['Leaderboard']
    ws_lb['B3'] = PERIOD_LABEL
    ws_lb['B3'].font = Font(name="Arial", bold=True, color="1565C0", size=11)
    ws_lb['B3'].fill = PatternFill("solid", fgColor="FFF9C4")

    wb.save(WORKBOOK)
    print(f"\nSaved: {WORKBOOK}")

    # ── SUMMARY ──
    total_orders = (sum(d['orders'] for d in ue_data.values()) +
                    sum(d['orders'] for d in dd_data.values()) +
                    sum(d['orders'] for d in gh_data.values()))
    total_ratings = (sum(d['ratings'] for d in ue_data.values()) +
                     sum(d['ratings'] for d in dd_data.values()) +
                     sum(d['ratings'] for d in gh_data.values()))
    print(f"\n{'='*50}")
    print(f"PERIOD: {PERIOD_LABEL}")
    print(f"STORES: {len(store_names)}")
    print(f"DATA ROWS: {data_rows} (UE + DD + GH)")
    print(f"TOTAL ORDERS: {total_orders:,}")
    print(f"TOTAL RATINGS: {total_ratings:,}")
    if not DD_REVIEWS:
        print(f"⚠️  DD 'Loved' ratings missing — need Customer Reviews export from Merchant Portal")
    print(f"{'='*50}")

if __name__ == '__main__':
    main()
