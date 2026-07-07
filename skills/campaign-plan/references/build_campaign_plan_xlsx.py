#!/usr/bin/env python3
"""Build a client Campaign Plan & Performance Tracker .xlsx.

Reads campaign rows from a tracker CSV (produced by the team from platform exports,
or exported from the Campaign Planning DB) and renders a formatted 3-tab workbook:
Dashboard (rollups overall / by platform / ads vs offers + chart) + Campaign Tracker
(conditional-formatted status) + Legend.

Usage:
  python build_campaign_plan_xlsx.py --client "<Client Name>" --tracker-csv rows.csv --output out.xlsx

Tracker CSV columns (header row, in this order):
  Campaign, Platform, Type, Offer/Ad Detail, Locations, Status, Days in Queue,
  Flight Start, Flight End, Target ROAS, Actual ROAS, Spend, Attributed Sales,
  Incremental Orders, In-Platform Campaign Name, Notes

Status values: Live | Proposed | Blocked-on-client | Ended
Type values: Offer | Ad
"""
from __future__ import annotations
import argparse
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

CHARCOAL, SAGE, RED, AMBER, BLUE, GRAY, WHITE = "1F2937", "84CC16", "B91C1C", "F59E0B", "2563EB", "9CA3AF", "FFFFFF"
TEAL = "0D9488"
thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_font = Font(name="Arial", size=10, bold=True, color=WHITE)
hdr_fill = PatternFill("solid", fgColor=CHARCOAL)
title_font = Font(name="Arial", size=16, bold=True, color=CHARCOAL)
sub_font = Font(name="Arial", size=9, italic=True, color="6B7280")
cell_font = Font(name="Arial", size=10, color=CHARCOAL)
kpi_label = Font(name="Arial", size=9, bold=True, color="6B7280")
kpi_value = Font(name="Arial", size=20, bold=True, color=CHARCOAL)
section_font = Font(name="Arial", size=12, bold=True, color=CHARCOAL)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Direct (static) per-cell styling for the Status column. Conditional-formatting RULES
# don't survive the xlsx -> Numbers / Google Sheets conversion (the fill drops, leaving
# white text invisible), so we paint the status cell directly. Survives everywhere.
STATUS_STYLE = {
    "Live": SAGE, "Approved": TEAL, "Proposed": BLUE,
    "Blocked-on-client": AMBER, "Ended": GRAY,
}

TRACKER_COLS = ["Campaign", "Platform", "Type", "Offer / Ad Detail", "Locations", "Segment",
                "Status", "Days in Queue", "Flight Start", "Flight End",
                "Target ROAS", "Actual ROAS", "Spend ($)", "Attributed Sales ($)",
                "Incremental Orders", "In-Platform Campaign Name", "Notes"]
SEGMENTS = ["All", "New", "Existing", "Lapsed"]  # client-facing customer segments

# Ad Performance tab: paid-placement funnel. Impressions/Clicks/Spend/Orders/Sales come
# from the platform Ads Manager exports; CTR/CPC/ROAS/CPO are computed live in-sheet.
ADS_COLS = ["Campaign", "Platform", "Locations", "Status", "Impressions", "Clicks", "CTR",
            "Spend ($)", "CPC ($)", "Orders", "Attributed Sales ($)", "ROAS", "CPO ($)"]
# Input CSV the team/feed provides (computed columns are derived, not supplied):
ADS_INPUT_COLS = ["Campaign", "Platform", "Locations", "Status",
                  "Impressions", "Clicks", "Spend", "Orders", "Attributed Sales"]

def _perf_num(s):
    """Parse a weekly-reporting formatted value ('$1,234', '6.2', '--', '--*') to float or None."""
    if s is None:
        return None
    s = str(s).strip().replace("$", "").replace(",", "").replace("x", "")
    if s in ("", "--", "--*", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _derive_type(campaign_type: str) -> str:
    """Map weekly-reporting 'Campaign Type' to the tracker's Ad/Offer dimension."""
    t = (campaign_type or "").lower()
    if any(k in t for k in ("sponsor", "featured", "ad", "paid", "listing")):
        return "Ad"
    return "Offer"


def _loc_tokens(locations: str) -> list[str]:
    """Split a tracker Locations cell ('San Jose / Pasadena') into normalized tokens."""
    raw = str(locations or "").replace(",", "/").split("/")
    return [t.strip().lower() for t in raw if t.strip()]


def _loc_match(tracker_locs: str, perf_loc: str) -> bool:
    """A perf row's single Location matches a tracker row if it overlaps any token,
    or the tracker row targets All/blank/whole-account."""
    toks = _loc_tokens(tracker_locs)
    pl = str(perf_loc or "").strip().lower()
    if not toks or "all" in toks:
        return True
    if not pl or pl in ("all", "total"):
        return True
    return any(t in pl or pl in t for t in toks)


# Tokens that don't identify a campaign — drop them before name-overlap matching.
_NAME_STOPWORDS = {"all", "the", "and", "new", "aggressive", "tweak", "test", "promo",
                   "depth", "launch", "ads", "ad", "offer", "listing", "sponsored", "paid"}


def _name_tokens(*parts: str) -> set:
    out = set()
    for part in parts:
        for raw in str(part or "").replace("|", " ").replace("/", " ").replace("(", " ").replace(")", " ").split():
            t = raw.strip().lower()
            if len(t) >= 3 and t not in _NAME_STOPWORDS:
                out.add(t)
    return out


def _name_match(tracker_campaign: str, tracker_inplatform: str, perf_campaign_type: str) -> bool:
    """Require a meaningful token overlap between the perf row's Campaign Type and the
    tracker row's Campaign name or In-Platform Campaign Name. Disambiguates concurrent
    live campaigns on the same platform+locations (e.g. Spend X Save Y vs Friday Depth)."""
    perf_toks = _name_tokens(perf_campaign_type)
    if not perf_toks:
        return True  # perf row has no identifying name; fall back to platform+type+location
    track_toks = _name_tokens(tracker_campaign, tracker_inplatform)
    return bool(perf_toks & track_toks)


def apply_campaign_perf(tracker_rows: list[list], perf_rows: list[dict], overwrite: bool = False):
    """Fold weekly-reporting campaign_performance.csv rows into the tracker's performance
    columns. Match key: Platform + derived Type + Location overlap + campaign-name token
    overlap, restricted to tracker rows whose Status is Live or Ended (Proposed/Blocked
    campaigns haven't run, so they get no performance). Sums all matching perf rows into
    each tracker row. Returns (matched_tracker_count, unmatched_perf_rows).

    Columns filled (0-indexed): 11 Actual ROAS, 12 Spend, 13 Attributed Sales, 14 Incremental Orders.
    Only fills empty cells unless overwrite=True. Never silently drops perf rows — anything that
    matched no tracker row is returned for the GM to map.
    """
    used = [False] * len(perf_rows)
    matched_tracker = 0
    for row in tracker_rows:
        platform, ttype, locs, status = row[1], row[2], row[4], row[6]
        if str(status).strip() not in ("Live", "Ended"):
            continue  # no performance for campaigns that haven't run
        spend = sales = orders = 0.0
        hit = False
        for i, p in enumerate(perf_rows):
            if str(p.get("Platform", "")).strip().lower() != str(platform).strip().lower():
                continue
            if _derive_type(p.get("Campaign Type", "")) != str(ttype).strip():
                continue
            if not _loc_match(locs, p.get("Location", "")):
                continue
            if not _name_match(row[0], row[15], p.get("Campaign Type", "")):
                continue
            sp, sa, od = _perf_num(p.get("Spend")), _perf_num(p.get("Sales")), _perf_num(p.get("Orders"))
            if sp is None and sa is None and od is None:
                continue
            spend += sp or 0.0
            sales += sa or 0.0
            orders += od or 0.0
            used[i] = True
            hit = True
        if not hit:
            continue
        matched_tracker += 1
        roas = round(sales / spend, 2) if spend else None
        fills = {12: round(spend, 2), 13: round(sales, 2), 14: int(round(orders)), 11: roas}
        for col, val in fills.items():
            if val is None:
                continue
            if overwrite or row[col] in ("", None):
                row[col] = val
    unmatched = [perf_rows[i] for i in range(len(perf_rows)) if not used[i]]
    return matched_tracker, unmatched


def read_perf_csv(path: str) -> list[dict]:
    with open(path, newline="") as f:
        return list(csv.DictReader(f))


def read_ads_csv(path: str) -> list[dict]:
    """Read the ads-detail CSV (ADS_INPUT_COLS). Numeric fields coerced to float/int."""
    out = []
    with open(path, newline="") as f:
        for r in csv.DictReader(f):
            for k in ("Impressions", "Clicks", "Spend", "Orders", "Attributed Sales"):
                v = (r.get(k) or "").replace("$", "").replace(",", "").strip()
                r[k] = float(v) if v not in ("", "--", "--*") else None
            out.append(r)
    return out


def read_tracker_csv(path: str) -> list[list]:
    rows = []
    with open(path, newline="") as f:
        reader = csv.reader(f)
        next(reader, None)  # header
        for row in reader:
            row = (row + [""] * len(TRACKER_COLS))[:len(TRACKER_COLS)]
            for i in (10, 11, 12, 13, 14):  # numeric-ish columns (Target/Actual ROAS, Spend, Sales, Orders)
                if row[i] not in ("", None):
                    try:
                        row[i] = float(row[i])
                    except (ValueError, TypeError):
                        pass
            rows.append(row)
    return rows


def build(client: str, tracker_rows: list[list], highlights: list[str], out: str, ads_rows=None):
    ads_rows = ads_rows or []
    wb = Workbook()
    tr = wb.active
    tr.title = "Campaign Tracker"
    for c, h in enumerate(TRACKER_COLS, start=1):
        cell = tr.cell(row=1, column=c, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = hdr_font, hdr_fill, center, border
    for r, row in enumerate(tracker_rows, start=2):
        for c, val in enumerate(row, start=1):
            cell = tr.cell(row=r, column=c, value=val)
            cell.font = cell_font
            cell.alignment = left if c in (1, 4, 5, 6, 16, 17) else center
            cell.border = border
            if c in (11, 12):
                cell.number_format = '0.0"x"'
            if c in (13, 14):
                cell.number_format = '$#,##0'
            if c == 7 and val in STATUS_STYLE:  # paint Status cell directly (survives Sheets/Numbers)
                cell.fill = PatternFill("solid", fgColor=STATUS_STYLE[val])
                cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    widths = [26, 11, 8, 28, 16, 11, 17, 9, 12, 11, 10, 10, 11, 14, 12, 26, 30]
    for i, w in enumerate(widths, start=1):
        tr.column_dimensions[get_column_letter(i)].width = w
    tr.freeze_panes = "A2"
    tr.row_dimensions[1].height = 30
    last = len(tracker_rows) + 1
    rng = f"G2:G{last}"
    tr.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Live"'], fill=PatternFill("solid", fgColor=SAGE), font=Font(color=WHITE, bold=True)))
    tr.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Approved"'], fill=PatternFill("solid", fgColor=TEAL), font=Font(color=WHITE, bold=True)))
    tr.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Proposed"'], fill=PatternFill("solid", fgColor=BLUE), font=Font(color=WHITE, bold=True)))
    tr.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Blocked-on-client"'], fill=PatternFill("solid", fgColor=AMBER), font=Font(color=WHITE, bold=True)))
    tr.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Ended"'], fill=PatternFill("solid", fgColor=GRAY), font=Font(color=WHITE, bold=True)))

    db = wb.create_sheet("Dashboard", 0)
    db.sheet_view.showGridLines = False
    db["B2"] = f"{client} | Campaign Performance Dashboard"; db["B2"].font = title_font
    db["B3"] = "Rolling 90-day window. Performance cells auto-populate from weekly platform exports."; db["B3"].font = sub_font
    db["B5"] = "Highlights this cycle"; db["B5"].font = section_font
    for i, h in enumerate(highlights):
        c = db.cell(row=6 + i, column=2, value="• " + h); c.font = cell_font; c.alignment = left
        db.merge_cells(start_row=6 + i, start_column=2, end_row=6 + i, end_column=8)
    base = 6 + len(highlights) + 1
    db.cell(row=base, column=2, value="Overall").font = section_font
    kpis = [("Campaigns Live", '=COUNTIF(\'Campaign Tracker\'!G:G,"Live")', None),
            ("Proposed", '=COUNTIF(\'Campaign Tracker\'!G:G,"Proposed")', None),
            ("Blocked", '=COUNTIF(\'Campaign Tracker\'!G:G,"Blocked-on-client")', None),
            ("Total Spend", "=SUM('Campaign Tracker'!M:M)", '$#,##0'),
            ("Blended ROAS", '=IFERROR(SUM(\'Campaign Tracker\'!N:N)/SUM(\'Campaign Tracker\'!M:M),"--")', '0.0"x"')]
    for i, (label, formula, fmt) in enumerate(kpis):
        col = 2 + i * 2
        lc = db.cell(row=base + 1, column=col, value=label); lc.font = kpi_label; lc.alignment = center
        vc = db.cell(row=base + 2, column=col, value=formula); vc.font = kpi_value; vc.alignment = center
        if fmt:
            vc.number_format = fmt
        db.merge_cells(start_row=base + 1, start_column=col, end_row=base + 1, end_column=col + 1)
        db.merge_cells(start_row=base + 2, start_column=col, end_row=base + 2, end_column=col + 1)
    pr = base + 5
    db.cell(row=pr, column=2, value="By Platform").font = section_font
    for c, h in enumerate(["Platform", "Spend ($)", "Attributed Sales ($)", "ROAS", "Incremental Orders"], start=2):
        cell = db.cell(row=pr + 1, column=c, value=h); cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = border
    platforms = ["Uber Eats", "DoorDash", "Grubhub"]
    for i, p in enumerate(platforms):
        r = pr + 2 + i
        db.cell(row=r, column=2, value=p).font = cell_font
        db.cell(row=r, column=3, value=f'=SUMIF(\'Campaign Tracker\'!B:B,"{p}",\'Campaign Tracker\'!M:M)').number_format = '$#,##0'
        db.cell(row=r, column=4, value=f'=SUMIF(\'Campaign Tracker\'!B:B,"{p}",\'Campaign Tracker\'!N:N)').number_format = '$#,##0'
        db.cell(row=r, column=5, value=f'=IFERROR(SUMIF(\'Campaign Tracker\'!B:B,"{p}",\'Campaign Tracker\'!N:N)/SUMIF(\'Campaign Tracker\'!B:B,"{p}",\'Campaign Tracker\'!M:M),"--")').number_format = '0.0"x"'
        db.cell(row=r, column=6, value=f'=SUMIF(\'Campaign Tracker\'!B:B,"{p}",\'Campaign Tracker\'!O:O)')
        for c in range(2, 7):
            db.cell(row=r, column=c).border = border
            db.cell(row=r, column=c).alignment = center if c > 2 else left

    # By Segment — New / Existing / Lapsed / All, summed from the Segment column (F).
    sg = pr + 2 + len(platforms) + 1
    db.cell(row=sg, column=2, value="By Segment").font = section_font
    for c, h in enumerate(["Segment", "Spend ($)", "Attributed Sales ($)", "ROAS", "Orders", "Count"], start=2):
        cell = db.cell(row=sg + 1, column=c, value=h); cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = border
    for i, s in enumerate(SEGMENTS):
        r = sg + 2 + i
        db.cell(row=r, column=2, value=s).font = cell_font
        db.cell(row=r, column=3, value=f'=SUMIF(\'Campaign Tracker\'!F:F,"{s}",\'Campaign Tracker\'!M:M)').number_format = '$#,##0'
        db.cell(row=r, column=4, value=f'=SUMIF(\'Campaign Tracker\'!F:F,"{s}",\'Campaign Tracker\'!N:N)').number_format = '$#,##0'
        db.cell(row=r, column=5, value=f'=IFERROR(SUMIF(\'Campaign Tracker\'!F:F,"{s}",\'Campaign Tracker\'!N:N)/SUMIF(\'Campaign Tracker\'!F:F,"{s}",\'Campaign Tracker\'!M:M),"--")').number_format = '0.0"x"'
        db.cell(row=r, column=6, value=f'=SUMIF(\'Campaign Tracker\'!F:F,"{s}",\'Campaign Tracker\'!O:O)')
        db.cell(row=r, column=7, value=f'=COUNTIF(\'Campaign Tracker\'!F:F,"{s}")')
        for c in range(2, 8):
            db.cell(row=r, column=c).border = border
            db.cell(row=r, column=c).alignment = center if c > 2 else left

    yr = sg + 2 + len(SEGMENTS) + 1
    db.cell(row=yr, column=2, value="Ads vs Offers").font = section_font
    for c, h in enumerate(["Type", "Spend ($)", "Attributed Sales ($)", "ROAS", "Count"], start=2):
        cell = db.cell(row=yr + 1, column=c, value=h); cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = border
    for i, t in enumerate(["Offer", "Ad"]):
        r = yr + 2 + i
        db.cell(row=r, column=2, value=t).font = cell_font
        if t == "Ad" and ads_rows:
            # Pull ad totals from the Ad Performance tab so this row agrees with the funnel.
            spend_f = "=SUM('Ad Performance'!H:H)"
            sales_f = "=SUM('Ad Performance'!K:K)"
            roas_f = "=IFERROR(SUM('Ad Performance'!K:K)/SUM('Ad Performance'!H:H),\"--\")"
            count_f = "=COUNTA('Ad Performance'!A4:A1000)"
        else:
            spend_f = f'=SUMIF(\'Campaign Tracker\'!C:C,"{t}",\'Campaign Tracker\'!M:M)'
            sales_f = f'=SUMIF(\'Campaign Tracker\'!C:C,"{t}",\'Campaign Tracker\'!N:N)'
            roas_f = f'=IFERROR(SUMIF(\'Campaign Tracker\'!C:C,"{t}",\'Campaign Tracker\'!N:N)/SUMIF(\'Campaign Tracker\'!C:C,"{t}",\'Campaign Tracker\'!M:M),"--")'
            count_f = f'=COUNTIF(\'Campaign Tracker\'!C:C,"{t}")'
        db.cell(row=r, column=3, value=spend_f).number_format = '$#,##0'
        db.cell(row=r, column=4, value=sales_f).number_format = '$#,##0'
        db.cell(row=r, column=5, value=roas_f).number_format = '0.0"x"'
        db.cell(row=r, column=6, value=count_f)
        for c in range(2, 7):
            db.cell(row=r, column=c).border = border
            db.cell(row=r, column=c).alignment = center if c > 2 else left
    # Ads funnel detail — totals across the Ad Performance tab. Only when ads data present.
    fr = yr + 5
    if ads_rows:
        db.cell(row=fr, column=2, value="Ads — Funnel Detail").font = section_font
        ad_metrics = [
            ("Impressions", "=SUM('Ad Performance'!E:E)", "#,##0"),
            ("Clicks", "=SUM('Ad Performance'!F:F)", "#,##0"),
            ("CTR", "=IFERROR(SUM('Ad Performance'!F:F)/SUM('Ad Performance'!E:E),\"--\")", "0.00%"),
            ("Ad Spend", "=SUM('Ad Performance'!H:H)", "$#,##0"),
            ("CPC", "=IFERROR(SUM('Ad Performance'!H:H)/SUM('Ad Performance'!F:F),\"--\")", "$0.00"),
            ("Ad Sales", "=SUM('Ad Performance'!K:K)", "$#,##0"),
            ("Ad ROAS", "=IFERROR(SUM('Ad Performance'!K:K)/SUM('Ad Performance'!H:H),\"--\")", '0.0"x"'),
        ]
        for c, (label, _, _) in enumerate(ad_metrics, start=2):
            cell = db.cell(row=fr + 1, column=c, value=label)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = border
        for c, (_, formula, fmt) in enumerate(ad_metrics, start=2):
            cell = db.cell(row=fr + 2, column=c, value=formula)
            cell.font = cell_font; cell.alignment = center; cell.border = border; cell.number_format = fmt
        db.cell(row=fr + 3, column=2,
                value="Funnel detail per ad campaign on the Ad Performance tab.").font = sub_font

    # Spend-by-platform chart, anchored BELOW all tables so nothing overlaps (overlapping
    # objects also convert unpredictably into Google Sheets).
    chart_row = (fr + 5) if ads_rows else (yr + 5)
    chart = BarChart(); chart.title = "Spend by Platform"; chart.type = "col"; chart.height = 8; chart.width = 18
    chart.add_data(Reference(db, min_col=3, min_row=pr + 1, max_row=pr + 1 + len(platforms)), titles_from_data=True)
    chart.set_categories(Reference(db, min_col=2, min_row=pr + 2, max_row=pr + 1 + len(platforms)))
    chart.legend = None
    db.add_chart(chart, f"B{chart_row}")

    for i, w in enumerate([3, 18, 16, 18, 12, 16, 14, 14], start=1):
        db.column_dimensions[get_column_letter(i)].width = w

    # ── Ad Performance tab (paid-placement funnel) ──────────────────────────
    if ads_rows:
        ap = wb.create_sheet("Ad Performance")
        ap.sheet_view.showGridLines = False
        ap["A1"] = f"{client} | Ad Performance (paid placements)"; ap["A1"].font = title_font
        ap.merge_cells("A1:M1")
        hr = 3
        for c, h in enumerate(ADS_COLS, start=1):
            cell = ap.cell(row=hr, column=c, value=h)
            cell.font, cell.fill, cell.alignment, cell.border = hdr_font, hdr_fill, center, border
        for i, a in enumerate(ads_rows):
            r = hr + 1 + i
            imp, clk, sp, od, sa = (a.get("Impressions"), a.get("Clicks"),
                                    a.get("Spend"), a.get("Orders"), a.get("Attributed Sales"))
            vals = [a.get("Campaign", ""), a.get("Platform", ""), a.get("Locations", ""), a.get("Status", ""),
                    imp, clk, f"=IFERROR(F{r}/E{r},\"--\")",
                    sp, f"=IFERROR(H{r}/F{r},\"--\")",
                    od, sa, f"=IFERROR(K{r}/H{r},\"--\")", f"=IFERROR(H{r}/J{r},\"--\")"]
            for c, v in enumerate(vals, start=1):
                cell = ap.cell(row=r, column=c, value=v)
                cell.font = cell_font
                cell.alignment = left if c in (1, 3) else center
                cell.border = border
            ap.cell(row=r, column=5).number_format = "#,##0"   # Impressions
            ap.cell(row=r, column=6).number_format = "#,##0"   # Clicks
            ap.cell(row=r, column=7).number_format = "0.00%"   # CTR
            ap.cell(row=r, column=8).number_format = "$#,##0"  # Spend
            ap.cell(row=r, column=9).number_format = "$0.00"   # CPC
            ap.cell(row=r, column=10).number_format = "#,##0"  # Orders
            ap.cell(row=r, column=11).number_format = "$#,##0" # Sales
            ap.cell(row=r, column=12).number_format = '0.0"x"' # ROAS
            ap.cell(row=r, column=13).number_format = "$0.00"  # CPO
            sc = ap.cell(row=r, column=4)  # Status pill
            if a.get("Status") in STATUS_STYLE:
                sc.fill = PatternFill("solid", fgColor=STATUS_STYLE[a["Status"]])
                sc.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        for col, w in zip("ABCDEFGHIJKLM", [26, 11, 16, 17, 12, 10, 8, 11, 9, 9, 14, 8, 9]):
            ap.column_dimensions[col].width = w
        ap.freeze_panes = "A4"
        ap.row_dimensions[hr].height = 28

    lg = wb.create_sheet("Legend & Cadence")
    lg.sheet_view.showGridLines = False
    lg["B2"] = "Legend & How We Work This"; lg["B2"].font = title_font
    legend = [("Status: Live", "Running now (approved + inside flight window)"),
              ("Status: Approved", "Signed off by you; scheduled, not yet started"),
              ("Status: Proposed", "Spice recommends; in our build/review pipeline, not yet sent for approval"),
              ("Status: Blocked-on-client", "Waiting on your team's sign-off (see Days in Queue)"),
              ("Status: Ended", "Completed or killed; outcome in Notes"),
              ("", ""),
              ("Type: Offer", "Merchant-funded promo (spend/save, BOGA, % off)"),
              ("Type: Ad", "Paid placement (sponsored listing, featured, paid social)"),
              ("", ""),
              ("Dashboard", "Auto-rolls up performance overall, by platform, ads vs offers from the Tracker"),
              ("Ad Performance tab", "Paid-placement funnel: impressions, clicks, CTR, CPC, ROAS, CPO per ad campaign"),
              ("Performance cells", "Populated weekly from UE/DD/GH platform exports. Empty = not yet pulled."),
              ("", ""),
              ("Cadence", "Updated every Friday. Blocked items need your sign-off. Tactical tweaks acknowledged 24h, shipped 48h unless they need your approval.")]
    for i, (k, v) in enumerate(legend):
        rk = lg.cell(row=4 + i, column=2, value=k); rk.font = Font(name="Arial", size=10, bold=True, color=CHARCOAL); rk.alignment = left
        rv = lg.cell(row=4 + i, column=3, value=v); rv.font = cell_font; rv.alignment = left
        lg.merge_cells(start_row=4 + i, start_column=3, end_row=4 + i, end_column=7)
    lg.column_dimensions["B"].width = 22
    lg.column_dimensions["C"].width = 62

    wb.save(out)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--client", required=True)
    ap.add_argument("--tracker-csv", required=True, help="CSV of campaign rows (see header spec), produced by refresh.py / db_to_tracker.py. Client-agnostic: no client data is embedded.")
    ap.add_argument("--campaign-perf-csv", default=None, help="weekly-reporting OUTPUT/campaign_performance.csv. Folds Spend/Sales/ROAS/Orders into matching tracker rows (no double-pull).")
    ap.add_argument("--overwrite-perf", action="store_true", help="Overwrite existing performance cells. Default fills only empty cells.")
    ap.add_argument("--ads-detail-csv", default=None, help="Ads funnel CSV (ADS_INPUT_COLS: Campaign,Platform,Locations,Status,Impressions,Clicks,Spend,Orders,Attributed Sales). Builds the Ad Performance tab + dashboard funnel.")
    ap.add_argument("--strict-unmatched", action="store_true", help="Surface ALL unmatched perf rows incl. Ad-type. Default suppresses Ad-type when ads_detail is provided (they live in Ad Performance, not Offer tracker).")
    ap.add_argument("--output", required=True)
    args = ap.parse_args()

    rows = read_tracker_csv(args.tracker_csv)
    highlights = ["Performance highlights populated from this cycle's results."]

    ads_rows = read_ads_csv(args.ads_detail_csv) if args.ads_detail_csv else []
    if ads_rows:
        print(f"ads detail: {len(ads_rows)} ad campaigns -> Ad Performance tab + dashboard funnel.")

    if args.campaign_perf_csv:
        perf = read_perf_csv(args.campaign_perf_csv)
        matched, unmatched = apply_campaign_perf(rows, perf, overwrite=args.overwrite_perf)
        print(f"campaign perf: {len(perf)} rows in, {matched} tracker rows updated.")
        # Ad-type unmatched rows are expected when ads_detail.csv carries them — those
        # live on the Ad Performance tab, not as Offer tracker rows. Suppress to reduce
        # recurring noise; --strict-unmatched surfaces them anyway.
        if ads_rows and not args.strict_unmatched:
            unmatched = [u for u in unmatched if _derive_type(u.get("Campaign Type","")) != "Ad"]
        if unmatched:
            print(f"⚠️  {len(unmatched)} perf rows matched NO tracker row — map or add as new campaigns:")
            for u in unmatched:
                print(f"    {u.get('Platform','?')} | {u.get('Campaign Type','?')} | {u.get('Location','?')} "
                      f"| spend {u.get('Spend','--')} | sales {u.get('Sales','--')}")

    out = build(args.client, rows, highlights, args.output, ads_rows=ads_rows)
    print(f"saved {out}: {os.path.getsize(out):,} bytes")


if __name__ == "__main__":
    main()
